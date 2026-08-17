"""
Read an intake workbook into a job dict and a list of unit rows.

The workbook is the input a person fills in -- Project tab for the job, Units
tab for the lineup -- and it is the only thing that changes between jobs. The
drafting conventions live in standard.json.

Both tabs are located by their header row rather than by a fixed row number, so
adding a note above the table, or another instruction line, does not silently
shift every field by one.

    from read_intake import read
    job, rows = read("intake_assignment.xlsx")
"""

from openpyxl import load_workbook


def _clean(value):
    """Cell -> stripped string. Blank, whitespace and None all become ''."""
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in ("none", "nan") else text


def _find_header(ws, wanted, limit=12):
    """Row index of the first row containing `wanted` in any cell."""
    for row in ws.iter_rows(min_row=1, max_row=limit):
        for cell in row:
            if _clean(cell.value).lower() == wanted:
                return cell.row
    raise ValueError(f"no header cell {wanted!r} in the first {limit} rows of {ws.title!r}")


def read_project(wb):
    """Project tab: a Field / Value table, returned as a plain dict.

    Fields left blank stay blank rather than being dropped, because a missing
    value is a thing preflight needs to report -- silently omitting the key
    would make an unanswered question indistinguishable from one never asked.
    """
    ws = wb["Project"]
    header = _find_header(ws, "field")
    job = {}
    for row in ws.iter_rows(min_row=header + 1, max_col=2):
        name = _clean(row[0].value)
        if name:
            job[name] = _clean(row[1].value) if len(row) > 1 else ""
    return job


def read_units(wb):
    """Units tab: one dict per unit, keyed by the header names.

    Rows with no unit number are skipped -- the template ships with blank rows
    below the example ones, and an empty row is not a cubicle.
    """
    ws = wb["Units"]
    header_row = _find_header(ws, "unit")
    headers = [_clean(c.value) for c in ws[header_row]]

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1):
        values = {h: _clean(c.value) for h, c in zip(headers, row) if h}
        if values.get("unit"):
            rows.append(values)
    return rows, [h for h in headers if h]


def read(path):
    """Return (job, rows, columns) for an intake workbook."""
    wb = load_workbook(path, data_only=True)
    missing = [tab for tab in ("Project", "Units") if tab not in wb.sheetnames]
    if missing:
        raise ValueError(f"{path}: missing tab(s) {', '.join(missing)}")

    job = read_project(wb)
    rows, columns = read_units(wb)
    if not rows:
        raise ValueError(f"{path}: the Units tab has no rows with a unit number")
    return job, rows, columns


if __name__ == "__main__":
    import sys

    job, rows, columns = read(sys.argv[1])
    filled = {k: v for k, v in job.items() if v}
    print(f"PROJECT  {len(filled)}/{len(job)} fields filled")
    for key in ("job", "title", "system_voltage", "bus_rating", "bus_ka_rating"):
        print(f"  {key:22} {job.get(key) or '(blank)'}")
    blank = [k for k, v in job.items() if not v]
    if blank:
        print(f"  blank: {', '.join(blank)}")

    print(f"\nUNITS  {len(rows)} rows, {len(columns)} columns")
    for r in rows:
        used = [c for c in columns if r.get(c) and c not in ("unit", "bus", "deck")]
        print(f"  unit {r['unit']:>3}  {r.get('tag', ''):<10} {len(used):>2} cells filled")
