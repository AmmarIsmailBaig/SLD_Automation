#!/usr/bin/env python3
"""
Fill intake_template.xlsx from a project data file and export the units CSV.

Keeps the template's formatting, dropdowns and header comments -- it only
writes values into the Project and Units tabs.

    python fill_intake.py assignment_data.json intake_assignment.xlsx assignment_units.csv
"""
import csv, json, sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

TEMPLATE = 'intake_template.xlsx'
MISSING = PatternFill('solid', fgColor='FDE7E7')   # not stated in the source document


def main(data_path, out_xlsx, out_csv):
    data = json.load(open(data_path))
    wb = load_workbook(TEMPLATE)

    # ---- Project tab: match on the field name in column A
    ws = wb['Project']
    ws.cell(1, 1, f"Project intake - {data['_source']}")
    for row in ws.iter_rows(min_row=5, max_col=1):
        name = row[0].value
        if name in data['project']:
            c = ws.cell(row[0].row, 2, data['project'][name])
            c.font = Font(name='Arial', size=10)
        elif name in data.get('project_missing', []):
            c = ws.cell(row[0].row, 2, '')
            c.fill = MISSING

    # ---- Units tab: header order drives the column mapping
    u = wb['Units']
    cols = [u.cell(1, i).value for i in range(1, u.max_column + 1) if u.cell(1, i).value]
    for i in range(1, u.max_column + 2):          # drop the example row
        u.cell(2, i).value = None

    for r, unit in enumerate(data['units'], start=2):
        for i, name in enumerate(cols, start=1):
            c = u.cell(r, i, unit.get(name, ''))
            c.font = Font(name='Arial', size=10)
            if name in unit.get('_missing', []):
                c.fill = MISSING

    wb.save(out_xlsx)

    # The workbook uses the engineers' word, 'schema_type'. build_sld.py and
    # sld_config.json use 'archetype'. Translate here so neither side has to change.
    csv_cols = [('archetype' if c == 'schema_type' else c) for c in cols]
    with open(out_csv, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=csv_cols, extrasaction='ignore')
        w.writeheader()
        for unit in data['units']:
            row = {c: unit.get(c, '') for c in cols}
            row['archetype'] = row.pop('schema_type', '')
            w.writerow(row)

    print(f'{out_xlsx}  ({len(data["units"])} units)')
    print(f'{out_csv}')
    miss = sorted({m for un in data['units'] for m in un.get('_missing', [])})
    if miss or data.get('project_missing'):
        print('\nnot stated in the source document:')
        for m in data.get('project_missing', []):
            print(f'   project.{m}')
        for m in miss:
            print(f'   units.{m}')


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2], sys.argv[3])
